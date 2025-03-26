import openpyxl
import json
import colors


class Converter:

    def __init__(self, results_path: str, file_path: str, branch: str, sem, subjects):
        self.results_path = results_path
        self.sem = sem
        self.branch = branch.upper()
        self.cols = int(subjects)
        if self.sem == '1':
            self.cols = 11  # for 1st year
        self.file_path = file_path
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.results = self.__load_results()
        self.__labels()

    def __load_results(self):
        try:
            with open(self.results_path, 'r') as file:
                res = json.loads(file.read())
                file.close()
        except FileNotFoundError:
            print(colors.red + "Save was Successful but File Failed to Read!\n"
                               "Please Re-Run The Script\n")
            input(colors.green + "Press Enter To Exit!")
            exit(1)
        results = sorted(res, key=lambda d: d.get('Pin') if d.get('Pin') is not None else '0')
        return results

    def __labels(self):
        row = 1
        self.sheet.cell(row=1, column=1).value = 'Pin No'
        sub = 1 + int(self.sem) * 100
        for col in range(2, self.cols + 2):
            self.sheet.cell(row=row, column=col).value = sub
            sub += 1
        self.sheet.cell(row=row, column=self.cols + 2).value = "Total"
        self.sheet.cell(row=row, column=self.cols + 3).value = "Result"

    def save_results(self):
        row = 2
        for res in self.results:
            if not res:
                continue
            lst = list(res.values())
            i = 0
            for col in range(1,
                             self.cols + 4):  # [self.cols + 4] is 1 for pin, 1 for total, 1 for result and 1 as truncater
                self.sheet.cell(row=row, column=col).value = lst[i]
                i += 1
            row += 1
        self.workbook.save(filename=self.file_path)


